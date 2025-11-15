from flask import Flask, request, jsonify
import openpyxl
import random
import re
import os
import logging
from datetime import datetime

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Путь к Excel-файлу с вопросами
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(BASE_DIR, "questions.xlsx")

if not os.path.exists(excel_path):
    raise FileNotFoundError(f"Файл {excel_path} не найден!")

# Загрузка Excel файла
workbook = openpyxl.load_workbook(excel_path)
sheet_names = workbook.sheetnames

# Словарь с ID картинок для Алисы
ALICE_IMAGE_IDS = {
  "6668": "1030494/57df06289997d0b975b0",
  "6667": "1652229/8c2f3fed1f1627a74e47",
  "6663": "1652229/baa39a8647b4f6a75fee",
  "6645": "13200873/bda9c083cfafae4764c0",
  "6638": "1652229/57167d8c94edef667ecd",
  "6581": "1656841/53b93a4f853f557f78c7",
  "6580": "13200873/7cb497753a3169364075",
  "6579": "1652229/510382cc5b64b80e5e2a",
  "6578": "1652229/ad19d477852d4983a8bb",
  "6577": "1652229/c86ca9b8bf5e9f90ec71",
  "6576": "1652229/cd6ae0622eae6676b4f7",
  "6450": "13200873/693416b28e1a692bdd35",
  "6448": "13200873/da4d8f3fb5560fbe2f83",
  "6447": "1652229/b5480c2992f320bd982c",
  "6445": "13200873/122e41c1471c376077a0",
  "6440": "1652229/7a11c4d2e5baa27cf694",
  "6439": "1652229/1e32128988a4d5cbc291",
  "6437": "13200873/4cff1e97b4a3874a6ab7",
  "6436": "1652229/70693d3d73372d9ed315",
  "6355": "1652229/8ecad1ff6204ef149c36",
  "6350": "13200873/130105b972bfaad7a6a2",
  "6341": "13200873/922fbc92c02908217f8f",
  "6339": "1652229/1adcfbf7b51b37afbf58",
  "6338": "1030494/0bdfef64f202a8434cdc",
  "6337": "13200873/94d7f9b538632030e7f7",
  "6336": "1652229/98ba45f1c700b4981529",
  "6335": "1652229/9625b4a017127d64b9be",
  "6334": "1652229/a2eda4bbaff5e65b906b",
  "6333": "1656841/9dfdb0d6c95992b5a951",
  "6331": "1030494/c99f14f3b415ff591720",
  "6328": "1656841/99c6872db80e0b98eb1c",
  "6327": "13200873/8f1fa4935070d2e61f5b",
  "6324": "1652229/2717d760e2051d3293ea",
  "6323": "13200873/45bdc3f99ff0e436599c",
  "6320": "1652229/04af5f550c859958f4b1",
  "6319": "1030494/15b814b429e95b27e74f",
  "6318": "1030494/cba0ca11424f30985538",
  "6317": "1030494/6bc1a313139bf200d22c",
  "6316": "1652229/bae40a955c89db0b5b0a",
  "6315": "13200873/a5d09861cae566a92117",
  "6312": "13200873/1956a369d6b86b1f7990",
  "6305": "13200873/fb8be027da02189590b4",
  "5040": "1030494/56b2f96bde63186754d3",
  "5039": "1030494/496003edbb33ab1f46d4",
  "5038": "1652229/aa801f79fb187d450782",
  "5037": "13200873/9da8fa5e416c26c272cc",
  "5036": "1030494/6ee58785fda8ff21f072",
  "5035": "1656841/68c3a48e11fae4cbb8dd",
  "5034": "1652229/6a225d83cba52471e617",
  "5033": "1030494/444f450e629883f6f278",
  "5032": "1652229/7db47257775052429ccc",
  "5031": "1652229/2fc3f4181c88b9c5ee34",
  "5030": "13200873/57a408f70e61e762dad4",
  "5029": "1030494/046ea5fe2a53b970e5b5",
  "5028": "1030494/69eceafa5248fb8cf5b7",
  "5027": "13200873/a5014401e7dbc21f48dd",
  "5026": "1652229/1de926c8b92c0e529c66",
  "5025": "1652229/5097a836778bd12184f5",
  "5024": "13200873/e57f7823fd4c657b89bd",
  "5023": "1652229/9c95d15c5deec031d9ab",
  "5022": "1652229/39ced67aacc92afe3d0f",
  "5021": "13200873/d71d3d4b040d953f6e54",
  "5020": "1652229/1fde3265a1fa92f140bb",
  "5019": "1652229/23252fc7e77fd5cabfe2",
  "5018": "13200873/a1f45a89a159e57df4c8",
  "5017": "13200873/3e394be0f661936986cd",
  "5016": "1652229/7cadae30862e5340386b",
  "5015": "1652229/bc50411764b9bad5aecc",
  "5014": "1030494/324a594712804beb1118",
  "5013": "13200873/8e4acb9daa710ea0ba5b",
  "5012": "1656841/424b1cc5cc3335e30633",
  "5011": "1030494/d250350f5e4da3ebf626",
  "5010": "13200873/6a01f2b046648df608d1",
  "5009": "1652229/60c39748f70d24114be0",
  "5008": "1652229/cbdbc728881f3188a674",
  "5007": "13200873/0c04ee5a29a10467409c",
  "5006": "1652229/3fa8d0501900dfc082bf",
  "5005": "1652229/fd3cd69ac7e41fc3a9cd",
  "5004": "13200873/e91ea27b6054533e4222",
  "5003": "1652229/29030b2d6211b28f4a8b",
  "5002": "13200873/2509fabbce9db3508444",
  "5001": "1652229/0588b20bf0145425d0cc",
  "4443": "1652229/7cd44a8537b52ca29719",
  "4441": "13200873/961049698548660c1edf",
  "2004": "1652229/bbb69d7330833e010ffc",
  "2001": "1652229/5ea053d30ac111cd4234",
  "1414": "13200873/337b2074f50590d411f3",
  "1413": "1030494/5ef86fa45ec54120e658",
  "1412": "1652229/fb02aae9dea10177a903",
  "1411": "1533899/87025db269e38a1082cd",
  "1410": "1652229/90531f188d67d353cf2d",
  "1409": "1652229/b5a9fd4aa93b935bdbba",
  "1408": "13200873/763b268abc5fc756ad2f",
  "1407": "13200873/65c3519909a8e2fbf86a",
  "1406": "1652229/70db0df5a8d8ee886664",
  "1405": "13200873/8f29be55086af4bc6999",
  "1404": "13200873/16e709ddb2a3ef1c0ad8",
  "1403": "13200873/5429645baaa05006b298",
  "1402": "1030494/36357db8ab33d200616c",
  "1401": "1652229/6f94e20db6dbdf7be122",
  "1202": "1652229/d69fdcf9558e6487d63c",
  "1201": "1652229/7245445e8d0d5ff8c31c",
  "1118": "13200873/91f492e641595f44bd23",
  "1117": "13200873/545b0087fe5c12cd5155",
  "1116": "1656841/71344d574c57fa13555f",
  "1115": "1533899/d7025c1a62fef9b1fd71",
  "1114": "13200873/ba8cce54d7f3d4ad71d0",
  "1113": "13200873/42ec11d4f472ce091bda",
  "1112": "1652229/bc11321dd775f5606aa2",
  "1111": "1652229/265d7e92d1185f55adf2"
}



def parse_options(options_str):
    """Парсинг строки с вариантами ответов"""
    if not options_str:
        return []
    return [opt.strip() for opt in str(options_str).split(';') if opt.strip()]


def parse_correct(correct_str):
    """Парсинг правильных ответов"""
    if not correct_str:
        return []
    matches = re.findall(r'([А-ЯЁA-Z]\))', str(correct_str))
    return matches


def get_alice_image_id(image_name):
    """Получение ID картинки для Алисы"""
    if not image_name:
        return None
    return ALICE_IMAGE_IDS.get(str(image_name).strip())


# Загрузка всех вопросов из Excel
quizzes = {}
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        question, options, correct, explanation, image = (row + (None, None, None, None, None))[:5]
        if not question:
            continue

        alice_image_id = get_alice_image_id(image)

        data.append({
            "Вопрос": str(question).strip(),
            "Варианты": parse_options(options),
            "Правильный": parse_correct(correct),
            "Пояснение": str(explanation).strip() if explanation else "",
            "Изображение": alice_image_id
        })
    quizzes[sheet_name] = data


def get_random_question(topic, previous_questions=None):
    """Получение случайного вопроса по теме"""
    if topic not in quizzes or not quizzes[topic]:
        return None

    if previous_questions is None:
        previous_questions = []

    available_questions = [q for q in quizzes[topic] if q["Вопрос"] not in previous_questions]

    if not available_questions:
        available_questions = quizzes[topic]

    return random.choice(available_questions)


def normalize_answer(user_answer):
    """Нормализация ответа пользователя"""
    if not user_answer:
        return ""

    user_answer = user_answer.strip().lower()

    digit_to_letter = {"1": "а", "2": "б", "3": "в", "4": "г", "5": "д", "6": "е"}
    if user_answer in digit_to_letter:
        return digit_to_letter[user_answer]

    user_answer = re.sub(r'[).\s,]', '', user_answer)

    if user_answer and user_answer[0] in 'абвгде':
        return user_answer[0]

    return ""


def normalize_correct_answers(correct_answers):
    """Нормализация правильных ответов"""
    normalized = []
    for answer in correct_answers:
        clean_answer = re.sub(r'[)\s]', '', answer).lower()
        if clean_answer and clean_answer[0] in 'абвгде':
            normalized.append(clean_answer[0])
    return normalized


def parse_multiple_answers(command):
    """Парсинг нескольких ответов из команды"""
    cleaned = re.sub(r'[.,;]', ' ', command.lower())
    answers = cleaned.split()

    normalized_answers = []
    valid_answers = set()

    for answer in answers:
        normalized = normalize_answer(answer)
        if normalized and normalized not in valid_answers:
            normalized_answers.append(normalized)
            valid_answers.add(normalized)

    return normalized_answers


# Хранилище сессий пользователей
user_sessions = {}


@app.route("/", methods=["POST"])
def main():
    """Основной обработчик запросов от Алисы"""
    try:
        req = request.json
        if not req:
            return jsonify_error("Пустой запрос")

        command = req["request"]["command"].strip().lower()
        session = req.get("session", {})
        session_id = session.get("session_id")

        logger.info(f"Запрос: команда='{command}', session_id={session_id}")

        user_state = user_sessions.get(session_id, {})

        response = {
            "version": req["version"],
            "session": req["session"],
            "response": {"end_session": False, "text": "", "buttons": []},
            "session_state": {}
        }

        # Обработка новой сессии
        if session.get("new", False):
            user_sessions[session_id] = {}
            buttons = [{"title": name} for name in sheet_names]
            response["response"]["text"] = "Привет! Выберите тему для тестирования:"
            response["response"]["buttons"] = buttons
            logger.info("Новая сессия")
            return jsonify(response)

        # Обработка команды возврата в меню
        if any(nav_cmd in command for nav_cmd in ["назад", "меню", "главная", "выход"]):
            user_sessions[session_id] = {}
            buttons = [{"title": name} for name in sheet_names]
            response["response"]["text"] = "Вы вернулись в главное меню. Выберите тему:"
            response["response"]["buttons"] = buttons
            logger.info("Возврат в меню")
            return jsonify(response)

        # Обработка пропуска вопроса
        if any(skip_cmd in command for skip_cmd in ["пропустить", "следующий", "дальше", "skip", "next"]):
            if user_state.get("mode") == "question" and user_state.get("topic"):
                topic = user_state["topic"]
                previous_questions = user_state.get("previous_questions", [])

                next_question = get_random_question(topic, previous_questions)
                if next_question:
                    options_text = "\n".join([f"{opt}" for opt in next_question["Варианты"]]) if next_question["Варианты"] else ""

                    if next_question["Изображение"]:
                        response["response"]["card"] = {
                            "type": "BigImage",
                            "image_id": next_question["Изображение"],
                            "title": f"Тема: {topic}",
                            "description": f"{next_question['Вопрос']}\n\n{options_text}"
                        }
                        response["response"]["text"] = f"Вопрос пропущен. Смотрите картинку с вопросом выше."
                    else:
                        response_text = f"Вопрос пропущен.\n\nТема: \"{topic}\"\n\n{next_question['Вопрос']}\n\n{options_text}"
                        if len(response_text) > 1000:
                            response_text = response_text[:997] + "..."
                        response["response"]["text"] = response_text

                    updated_previous_questions = previous_questions + [next_question["Вопрос"]]
                    user_sessions[session_id] = {
                        "topic": topic,
                        "question": next_question,
                        "previous_questions": updated_previous_questions,
                        "mode": "question"
                    }
                else:
                    response["response"]["text"] = "Вопросы в этой теме закончились."
                    user_sessions[session_id] = {}

                response["response"]["buttons"] = [
                    {"title": "Пропустить"},
                    {"title": "Назад в меню"}
                ]
                logger.info("Вопрос пропущен")
                return jsonify(response)

        # Обработка команды помощи
        if command in ["помощь", "help", "что делать", "правила"]:
            if user_state.get("mode") == "question":
                response["response"]["text"] = f"Вы в режиме вопроса по теме '{user_state['topic']}'. Произнесите номер ответа (1-6) или букву (А-Е). Можно несколько ответов через пробел. Скажите 'пропустить' для перехода к следующему вопросу. Или скажите 'назад' для возврата в меню."
            else:
                response["response"]["text"] = "Я помогу вам подготовиться к экзамену. Выберите тему для тестирования или скажите 'назад' в любой момент. Во время тестирования можно пропускать вопросы командой 'пропустить'."
            response["response"]["buttons"] = [{"title": "Назад в меню"}]
            logger.info("Показана помощь")
            return jsonify(response)

        # Выбор темы
        # Выбор темы
        for sheet_name in sheet_names:
            # Проверяем все варианты названий
            if (command == sheet_name.lower() or
                    (sheet_name == "Первая помощь" and command in ["1 помощь", "первая помощь", "1помощь",
                                                                   "перваяпомощь"])):

                topic = sheet_name
                question = get_random_question(topic)
                if not question:
                    response["response"]["text"] = f"В теме '{topic}' нет вопросов."
                    response["response"]["buttons"] = [{"title": "Назад в меню"}]
                    logger.warning(f"В теме '{topic}' нет вопросов")
                    return jsonify(response)

                options_text = "\n".join([f"{opt}" for opt in question["Варианты"]]) if question["Варианты"] else ""

                if question["Изображение"]:
                    response["response"]["card"] = {
                        "type": "BigImage",
                        "image_id": question["Изображение"],
                        "title": f"Тема: {topic}",
                        "description": f"{question['Вопрос']}\n\n{options_text}"
                    }
                    response["response"]["text"] = f"Смотрите вопрос на картинке. {question['Вопрос']}"
                else:
                    response_text = f'Тема: "{topic}"\n\n{question["Вопрос"]}\n\n{options_text}'
                    if len(response_text) > 1000:
                        response_text = response_text[:997] + "..."
                    response["response"]["text"] = response_text

                response["response"]["buttons"] = [
                    {"title": "Пропустить"},
                    {"title": "Назад в меню"}
                ]

                user_sessions[session_id] = {
                    "topic": topic,
                    "question": question,
                    "previous_questions": [question["Вопрос"]],
                    "mode": "question"
                }

                logger.info(f"Выбрана тема '{topic}'")
                return jsonify(response)

        # Обработка ответа на вопрос
        if user_state.get("mode") == "question" and user_state.get("topic") and user_state.get("question"):
            topic = user_state["topic"]
            current_question = user_state["question"]
            previous_questions = user_state.get("previous_questions", [])

            logger.info(f"Обработка ответа для темы '{topic}': '{command}'")

            user_answers = parse_multiple_answers(command)
            correct_answers_normalized = normalize_correct_answers(current_question["Правильный"])

            logger.info(f"Ответы пользователя: {user_answers}")
            logger.info(f"Правильные ответы: {correct_answers_normalized}")

            if not user_answers:
                response["response"]["text"] = f"Не понял ответ '{command}'. Используйте цифры 1-6 или буквы А-Е. Пример: '1', 'а', '1 2', 'а б'. Скажите 'пропустить' для перехода к следующему вопросу. Или скажите 'назад' для возврата в меню."
                response["response"]["buttons"] = [
                    {"title": "Пропустить"},
                    {"title": "Назад в меню"}
                ]
                user_sessions[session_id] = user_state
                return jsonify(response)

            correct_given = [ans for ans in user_answers if ans in correct_answers_normalized]
            incorrect_given = [ans for ans in user_answers if ans not in correct_answers_normalized]

            if not incorrect_given and len(correct_given) == len(correct_answers_normalized):
                text = "Верно!"
            elif not incorrect_given and len(correct_given) > 0:
                missing = [ans for ans in correct_answers_normalized if ans not in user_answers]
                missing_text = ", ".join([f"{ans.upper()})" for ans in missing])
                text = f"Частично верно! Вы выбрали правильные ответы, но не хватает: {missing_text}\n\n{current_question['Пояснение']}"
            elif len(correct_given) > 0 and len(incorrect_given) > 0:
                correct_text = ", ".join([f"{ans.upper()})" for ans in correct_given])
                incorrect_text = ", ".join([f"{ans.upper()})" for ans in incorrect_given])
                text = f"Частично верно! Правильные: {correct_text}, неправильные: {incorrect_text}\n\n{current_question['Пояснение']}"
            else:
                correct_text = ", ".join(current_question["Правильный"])
                text = f"Неверно.\nПравильный ответ: {correct_text}\n\n{current_question['Пояснение']}"

            next_question = get_random_question(topic, previous_questions)
            if next_question:
                options_text = "\n".join([f"{opt}" for opt in next_question["Варианты"]]) if next_question["Варианты"] else ""

                if next_question["Изображение"]:
                    response["response"]["card"] = {
                        "type": "BigImage",
                        "image_id": next_question["Изображение"],
                        "title": f"Тема: {topic}",
                        "description": f"{next_question['Вопрос']}\n\n{options_text}"
                    }
                    text += f"\n\nСледующий вопрос: смотрите на картинке выше."
                else:
                    text += f"\n\nСледующий вопрос:\n{next_question['Вопрос']}\n\n{options_text}"

                if len(text) > 1000:
                    text = text[:997] + "..."

                updated_previous_questions = previous_questions + [next_question["Вопрос"]]
                user_sessions[session_id] = {
                    "topic": topic,
                    "question": next_question,
                    "previous_questions": updated_previous_questions,
                    "mode": "question"
                }
            else:
                text += "\n\nВопросы в этой теме закончились."
                user_sessions[session_id] = {}

            response["response"]["text"] = text
            response["response"]["buttons"] = [
                {"title": "Пропустить"},
                {"title": "Назад в меню"}
            ]
            return jsonify(response)

        # Команда не распознана
        buttons = [{"title": name} for name in sheet_names]
        response["response"]["text"] = "Пожалуйста, выберите тему из предложенных ниже."
        response["response"]["buttons"] = buttons
        return jsonify(response)

    except Exception as e:
        logger.error(f"Ошибка обработки запроса: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify_error("Произошла ошибка. Пожалуйста, попробуйте еще раз.")


def jsonify_error(message):
    """Формирование ответа с ошибкой"""
    return jsonify({
        "version": "1.0",
        "response": {"text": message, "end_session": False},
        "session_state": {}
    })


@app.route("/", methods=["GET"])
def home():
    """Обработчик GET запросов для проверки работы сервера"""
    return jsonify({
        "status": "success",
        "message": "Навык Алисы работает.",
        "active_sessions": len(user_sessions),
        "topics_loaded": list(quizzes.keys())
    })

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    logger.info(f"Запуск сервера на порту {port}")
    app.run(host="0.0.0.0", port=port, debug=False)
